#import openpyxl as px
import pandas as pd
import Dataclasses as dc
import Fredon_projects_fields as pf
import streamlit as st
import datetime
import openpyxl as px
from openpyxl import Workbook

def get_data_columns(df):
    return df.columns.tolist()

def remove_columns(df, columns_to_remove):
    return df.drop(columns=columns_to_remove)

def get_list_of_projects(df):
    # Get the column name by index (column 1 = PROJECT_ID)
    project_col = df.columns[pf.PROJECT_ID] if pf.PROJECT_ID < len(df.columns) else None
    
    if project_col is None:
        print(f"Column index {pf.PROJECT_ID} not found in DataFrame")
        return []
    
    df_clean = df.dropna(subset=[project_col])
    df_clean = df_clean[df_clean[project_col].astype(str).str.strip() != ""]
    return df_clean[project_col].unique().tolist()

def get_projects_with_highest_complete(df):
    # Get column names by index
    projectid_col = df.columns[pf.PROJECT_ID] if pf.PROJECT_ID < len(df.columns) else None
    complete_col = df.columns[pf.PERCENTAGE_COMPLETE] if pf.PERCENTAGE_COMPLETE < len(df.columns) else None
    project_col = df.columns[pf.PROJECT_NAME] if pf.PROJECT_NAME < len(df.columns) else None
    
    if projectid_col is None or complete_col is None or project_col is None:
        st.write(f"Missing columns: PROJECT_ID={projectid_col}, COMPLETE={complete_col}, PROJECT_NAME={project_col}")
        return []
    
    # Clean the data
    df_clean = df.dropna(subset=[projectid_col, project_col, complete_col])
    
    # Convert %Complete to numeric if it's not already
    if df_clean[complete_col].dtype == 'object':
        # Remove % signs and convert to float
        df_clean = df_clean.copy()
        df_clean[complete_col] = df_clean[complete_col].astype(str).str.replace('%', '').str.replace(',', '').str.strip()
        df_clean[complete_col] = pd.to_numeric(df_clean[complete_col], errors='coerce')
        df_clean = df_clean.dropna(subset=[complete_col])
    
    # If values are between 0-1 (like 0.95), convert to percentage (like 95)
    max_val = df_clean[complete_col].max()
    
    if max_val <= 1.1:  # Allow for slight floating point errors above 1.0
        # Create new DataFrame with converted percentages
        data_dict = {
            projectid_col: df_clean[projectid_col].values,
            project_col: df_clean[project_col].values,
            complete_col: df_clean[complete_col].values * 100
        }
        df_clean = pd.DataFrame(data_dict)
    
    if df_clean.empty:
        st.write("No data remaining after cleaning")
        return []
    
    # Group by PROJECT_ID and get the maximum % Complete for each project
    # First, get the actual maximum values per group
    max_percentages = df_clean.groupby(projectid_col)[complete_col].max().reset_index()
    
    # Then get the corresponding project names by merging with the original data
    # Get one row per project (take the first occurrence for project name)
    project_info = df_clean.groupby(projectid_col)[project_col].first().reset_index()
    
    # Combine the data
    result = pd.merge(max_percentages, project_info, on=projectid_col)
    
    # Add status column - use 95 since we converted to percentage scale
    result["Status"] = result[complete_col].apply(lambda x: "Calibrate" if x > 95 else "Operational")
    
    # Format % Complete as a percentage string
    result["% Complete Formatted"] = result[complete_col].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")
    
    # Return as list of tuples: (Project_ID, Project_Name, Max_%, Status)
    projects = []
    for _, row in result.iterrows():
        projects.append((row[projectid_col], row[project_col], row["% Complete Formatted"], row["Status"]))

    return projects

def create_data_objects(df):
    project_names = get_list_of_projects(df)
    
    if not project_names:
        st.write("No project names found")
        return pd.DataFrame()
    
    # Get the project column name by index
    project_col = df.columns[pf.PROJECT_ID] if pf.PROJECT_ID < len(df.columns) else None
    
    if project_col is None:
        st.write(f"Column index {pf.PROJECT_ID} not found in DataFrame")
        return pd.DataFrame()
    
    # Create portfolio to hold all projects
    Portfolio = dc.Portfolio()
    all_project_data = pd.DataFrame()
    
    st.write(f"Processing {len(project_names)} projects...")
    
    # Loop through all project names
    for project_idx, project_name in enumerate(project_names):
        st.write(f"Processing project {project_idx + 1}/{len(project_names)}: {project_name}")
        
        # Filter data for current project
        project_data = df[df[project_col] == project_name]
        
        if project_data.empty:
            st.write(f"No data found for project: {project_name}")
            continue
        
        try:
            project_object = dc.Projects(
                Project_Name=project_data.iloc[0, pf.PROJECT_NAME],
                Project_ID=project_data.iloc[0, pf.PROJECT_ID] if pf.PROJECT_ID < len(df.columns) else "Unknown",
                Location=project_data.iloc[0, pf.LOCATION] if pf.LOCATION and pf.LOCATION < len(df.columns) else "Unknown",
                Post_Code=project_data.iloc[0, pf.POST_CODE] if pf.POST_CODE and pf.POST_CODE < len(df.columns) else "",
                Sector=project_data.iloc[0, pf.CLIENT_SECTOR] if pf.CLIENT_SECTOR < len(df.columns) else "Unknown",
                Portfolio_Bus_Unit_Dept_ID=project_data.iloc[0, pf.PORTFOLIO_BUS_UNIT_DEPT_ID] if pf.PORTFOLIO_BUS_UNIT_DEPT_ID < len(df.columns) else "Unknown",
                Asset_Type=project_data.iloc[0, pf.ASSET_TYPE] if pf.ASSET_TYPE and pf.ASSET_TYPE < len(df.columns) else "Other",
                Contract_Type=project_data.iloc[0, pf.CONTRACT_TYPE] if pf.CONTRACT_TYPE and pf.CONTRACT_TYPE < len(df.columns) else "Unknown",
                Contract_Financial=project_data.iloc[0, pf.CONTRACT_FINANCIAL] if pf.CONTRACT_FINANCIAL < len(df.columns) else "Unknown",
                Client=project_data.iloc[0, pf.CLIENT] if pf.CLIENT < len(df.columns) else "Unknown",
                Stage_of_Work=project_data.iloc[0, pf.STAGE_OF_WORK] if pf.STAGE_OF_WORK < len(df.columns) else "Unknown",
                Template_Version="0.2",
                Status="Operational",
                Comments=str(project_data.iloc[0, pf.COMMENTS]) if hasattr(pf, 'COMMENTS') and pf.COMMENTS and pf.COMMENTS < len(df.columns) else None
            )
            
            # Process monthly data for this project
            number_of_months = project_data.shape[0]
            for mon in range(number_of_months):
                try:
                    # Handle date conversion properly
                    if pf.DATE and pf.DATE < len(df.columns):
                        date_value = project_data.iloc[mon, pf.DATE]
                        if pd.isna(date_value):
                            record_date = ""  # Empty string if null
                        else:
                            # Convert pandas Timestamp to formatted string
                            if hasattr(date_value, 'date'):
                                record_date = date_value.strftime("%d/%m/%Y")  # Format as dd/mm/yyyy
                            elif isinstance(date_value, str):
                                # Handle string format like "2019-11-30 00:00:00"
                                try:
                                    parsed_date = pd.to_datetime(date_value)
                                    record_date = parsed_date.strftime("%d/%m/%Y")  # Format as dd/mm/yyyy
                                except:
                                    record_date = ""  # Empty string if parsing fails
                            else:
                                record_date = ""  # Empty string for other types
                    else:
                        record_date = ""  # Empty string if no date column
                    
                    monthly_record = dc.MonthlyRecord(
                        Date=record_date,
                        approved_budget=float(project_data.iloc[mon, pf.APPROVED_BUDGET]) if pf.APPROVED_BUDGET and pf.APPROVED_BUDGET < len(df.columns) else 0.0,
                        forecast_end_date=float(project_data.iloc[mon, pf.FORECAST_END_DATE]) if pf.FORECAST_END_DATE and pf.FORECAST_END_DATE < len(df.columns) else 0.0,
                        forecast_final_cost=float(project_data.iloc[mon, pf.FORECAST_FINAL_COST]) if pf.FORECAST_FINAL_COST and pf.FORECAST_FINAL_COST < len(df.columns) else 0.0,
                        contingency_remaining=float(project_data.iloc[mon, pf.CONTINGENCY_REMAINING]) if pf.CONTINGENCY_REMAINING and pf.CONTINGENCY_REMAINING < len(df.columns) else 0.0,
                        actual_cost_to_date=float(project_data.iloc[mon, pf.ACTUAL_COST_TO_DATE]) if pf.ACTUAL_COST_TO_DATE and pf.ACTUAL_COST_TO_DATE < len(df.columns) else 0.0,
                        forecast_final_revenue=float(project_data.iloc[mon, pf.FORECAST_FINAL_REVENUE]) if pf.FORECAST_FINAL_REVENUE and pf.FORECAST_FINAL_REVENUE < len(df.columns) else 0.0,
                        actual_revenue_to_date=float(project_data.iloc[mon, pf.ACTUAL_REVENUE_TO_DATE]) if pf.ACTUAL_REVENUE_TO_DATE and pf.ACTUAL_REVENUE_TO_DATE < len(df.columns) else 0.0,
                        notes=f"{float(project_data.iloc[mon, pf.PERCENTAGE_COMPLETE]) * 100:.0f}% Cost Complete" if pf.PERCENTAGE_COMPLETE and pf.PERCENTAGE_COMPLETE < len(df.columns) and pd.notnull(project_data.iloc[mon, pf.PERCENTAGE_COMPLETE]) else None
                       
                    )
                    project_object.Monthly_data.append(monthly_record)
                except (ValueError, IndexError, TypeError) as e:
                    st.write(f"Error processing month {mon+1} for project {project_name}: {e}")
                    continue
            
            # Add project to portfolio
            Portfolio.add_project(project_object)
            
            # Add column numbers to project data and append to all_project_data
            numbered_columns = {col: f"Col {i+1}: {col}" for i, col in enumerate(project_data.columns)}
            project_data_numbered = project_data.rename(columns=numbered_columns)
            all_project_data = pd.concat([all_project_data, project_data_numbered], ignore_index=True)
            
        except Exception as e:
            st.write(f"Error processing project {project_name}: {e}")
            continue
    
    st.write(f"Successfully processed {len(Portfolio.projects)} projects")
    
    # Display project summary
    st.write("Portfolio Summary:")
    st.data_editor(Portfolio.to_dataframe(), use_container_width=True)
    
    # Display monthly data
    st.write("All Monthly Data:")
    st.data_editor(Portfolio.monthly_data_to_dataframe(), use_container_width=True)
    
    # Add Excel export button
    if st.button("Export Portfolio to Excel"):
        created_files = create_excel_file_with_portfolio_data(Portfolio)
        
        if created_files:
            st.success(f"Excel files created for {len(Portfolio.projects)} projects:")
            for filename in created_files:
                st.write(f"- {filename}")
            
            # Offer download for each file
            for filename in created_files:
                with open(filename, "rb") as file:
                    file_label = "Calibrate Projects" if "Calibrate" in filename else "Operational Projects"
                    st.download_button(
                        label=f"Download {file_label}",
                        data=file.read(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{filename}"  # Unique key for each button
                    )
        else:
            st.warning("No files were created. Check if there are projects in the portfolio.")
    
    return all_project_data


  
    
    
def create_excel_file_with_portfolio_data(portfolio, base_filename="portfolio_output"):
    """
    Create two separate Excel files: one for Calibrate projects and one for Operational projects
    """
    # Separate projects by status
    calibrate_projects = []
    operational_projects = []
    
    for project in portfolio.projects:
        # Determine project status based on max percentage complete
        max_percentage = 0
        for monthly_record in project.Monthly_data:
            if monthly_record.notes and "% Cost Complete" in monthly_record.notes:
                try:
                    # Extract percentage from notes like "75% Cost Complete"
                    percentage_str = monthly_record.notes.split('%')[0]
                    percentage = float(percentage_str)
                    max_percentage = max(max_percentage, percentage)
                except:
                    continue
        
        # Classify project based on max percentage
        if max_percentage > 95:
            calibrate_projects.append(project)
        else:
            operational_projects.append(project)
    
    created_files = []
    
    # Create Calibrate projects file
    if calibrate_projects:
        calibrate_filename = f"{base_filename}_Calibrate.xlsx"
        _create_workbook_for_projects(calibrate_projects, calibrate_filename)
        
        # Validate formatting
        success, message = validate_excel_formatting(calibrate_filename)
        print(f"Calibrate file formatting validation: {message}")
        
        created_files.append(calibrate_filename)
    
    # Create Operational projects file
    if operational_projects:
        operational_filename = f"{base_filename}_Operational.xlsx"
        _create_workbook_for_projects(operational_projects, operational_filename)
        
        # Validate formatting
        success, message = validate_excel_formatting(operational_filename)
        print(f"Operational file formatting validation: {message}")
        
        created_files.append(operational_filename)
    
    return created_files

def _create_workbook_for_projects(projects, filename):
    """
    Helper function to create an Excel workbook for a list of projects with proper formatting
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    
    # Create a new workbook
    workbook = px.Workbook()
    
    # Remove the default sheet
    workbook.remove(workbook.active)
    
    # Define border styles to match reference template
    thin_border = Side(border_style="thin", color="000000")
    medium_border = Side(border_style="medium", color="000000")
    
    # Define styles to match the reference template
    header_font = Font(name='Calibri', size=18, bold=True)
    header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(top=medium_border, bottom=medium_border)
    
    summary_font = Font(name='Calibri', size=16, bold=True)
    summary_fill = PatternFill(start_color='FFAFE7E7', end_color='FFAFE7E7', fill_type='solid')
    summary_alignment = Alignment(horizontal='center', vertical='center')
    
    label_font = Font(name='Calibri', size=11, bold=True)
    value_font = Font(name='Calibri', size=11, bold=False)
    value_alignment = Alignment(horizontal='left', vertical='center')
    label_font_color = "2C8F8F"  # Dark teal color for labels
    
    version_font = Font(name='Calibri', size=10, bold=True)
    version_fill = PatternFill(start_color='FFAFE7E7', end_color='FFAFE7E7', fill_type='solid')
    version_alignment = Alignment(horizontal='right', vertical='center')
    version_border = Border(left=medium_border, top=medium_border)
    
    monthly_header_font = Font(name='Calibri', size=10, bold=True)
    monthly_header_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    monthly_header_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # Data row styles
    data_font = Font(name='Calibri', size=11, bold=False)
    data_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    # Create a worksheet for each project
    for idx, project_object in enumerate(projects):
        # Create worksheet with incremental naming: Project (1), Project (2), etc.
        sheet_name = f"Project ({idx + 1})"
        
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Set column widths to match reference template
        sheet.column_dimensions['A'].width = 12.55
        sheet.column_dimensions['B'].width = 13.0
        sheet.column_dimensions['C'].width = 13.0
        sheet.column_dimensions['D'].width = 13.0
        sheet.column_dimensions['E'].width = 13.0
        sheet.column_dimensions['F'].width = 13.0
        sheet.column_dimensions['G'].width = 13.0
        sheet.column_dimensions['H'].width = 13.0
        sheet.column_dimensions['I'].width = 7.44
        sheet.column_dimensions['J'].width = 31.44
        sheet.column_dimensions['K'].width = 0.63
        
        #Set row heights to match reference template
        sheet.row_dimensions[1].height = 45.0  # Header row height
        sheet.row_dimensions[2].height = 40.5  # Second row height
        sheet.row_dimensions[3].height = 9
        sheet.row_dimensions[20].height = 40.5  # Project summary labels
        sheet.row_dimensions[21].height = 75  # Monthly data headers
        
        # Template headers with formatting
        sheet["A1"] = "TEMPLATE VERSION:"
        sheet["A1"].font = version_font
       # sheet["A1"].fill = version_fill
        sheet["A1"].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        #sheet["A1"].border = version_border
        # Loop over a range of cells
    
        sheet["B1"] = project_object.Template_Version
        sheet["B1"].font = Font(name='Calibri', size=12, bold=True)
       # sheet["B1"].fill = version_fill
        sheet["B1"].alignment = Alignment(horizontal='center', vertical='center')
       # sheet["B1"].border = Border(top=medium_border)
        
        sheet["H1"] = "OCTANT BUSINESS - PROJECT DATA UPLOAD TEMPLATE"
        sheet["H1"].font = header_font
       # sheet["H1"].fill = header_fill
        sheet["H1"].alignment = header_alignment
       # sheet["H1"].border = header_border
        for row in sheet['A1:K1']:
            for cell in row:
                cell.fill = header_fill
                # Apply specific borders based on column
                left_border = medium_border if cell.column_letter in ['A', 'D'] else None
                right_border = medium_border if cell.column_letter == 'K' else None
                cell.border = Border(
                    top=medium_border, 
                    bottom=medium_border, 
                    left=left_border, 
                    right=right_border
                )
        sheet["F2"] = "PROJECT SUMMARY"
        sheet["F2"].font = summary_font
        sheet["F2"].fill = summary_fill
        sheet["F2"].alignment = summary_alignment       
        for row in sheet['A2:K2']:
            for cell in row:
                cell.fill = version_fill
                # Apply specific borders based on column
                left_border = medium_border if cell.column_letter == 'A' else None
                right_border = medium_border if cell.column_letter == 'K' else None
                cell.border = Border(
                    top=medium_border, 
                    bottom=medium_border, 
                    left=left_border, 
                    right=right_border        
                )
        #create borders for the summary rows, and data rows        
        for row in sheet['C4:E4']:
            for cell in row:
                cell.border = Border(
                    left=medium_border if cell.column_letter == 'C' else None, 
                    right=medium_border if cell.column_letter == 'E' else None, 
                    top=medium_border, 
                    bottom=medium_border
                )
        for row in sheet['C5:C9']:
            for cell in row:
                cell.border = Border(
                    left=medium_border, 
                    right=medium_border, 
                    top=medium_border, 
                    bottom=medium_border
                )
        for row in sheet['J4:J8']:
            for cell in row:
                cell.border = Border(
                    left=medium_border , 
                    right=medium_border, 
                    top=medium_border, 
                    bottom=medium_border
                )
        for row in sheet['C11:J12']:
            for cell in row:
                cell.border = Border(
                    left=medium_border if cell.column_letter == 'C' else None, 
                    right=medium_border if cell.column_letter == 'J' else None, 
                    top=medium_border, 
                    bottom=medium_border
                )
        for row in sheet['A14:F14']:
            for cell in row:
                cell.border = Border(
                    left=medium_border if cell.column_letter == 'A' else None, 
                    right=medium_border if cell.column_letter == 'F' else None, 
                    top=medium_border, 
                    bottom=medium_border
                )
        for row in sheet['A15:F19']:
            for cell in row:
                cell.fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
                cell.border = Border(
                    left=medium_border if cell.column_letter == 'A' else None, 
                    right=medium_border if cell.column_letter == 'F' else None, 
                    top=medium_border if cell.row == 15 else None, 
                    bottom=medium_border if cell.row == 19 else None
                )
        for row in sheet['A20:K20']:
            for cell in row:
                cell.fill = summary_fill
                cell.border = Border(
                    left=medium_border if cell.column_letter == 'A' else None, 
                    right=medium_border if cell.column_letter == 'K' else None, 
                    top=medium_border, 
                    bottom=medium_border
                )
        for row in sheet['K3:K21']:
            for cell in row:
                cell.border = Border(   
                    right=medium_border, 
                )
        # Project summary labels with formatting and borders
        labels = [
            ("A4", "Project Name*"),
            ("A5", "Project ID*"),
            ("A6", "Location*"),
            ("A7", "Post Code"),
            ("A8", "Sector*"),
            ("A9", "Portfolio/ Bus Unit/ Dept ID*"),
            ("A11", "Comments this period"),
            ("A12", "(max 200 Characters )"),
            ("H4", "Asset Type*"),
            ("H5", "Contract Type"),
            ("H6", "Contract Financial"),
            ("H7", "Client*"),
            ("H8", "Stage of Work*")
        ]
        
        for cell_ref, label_text in labels:
            sheet[cell_ref] = label_text
            sheet[cell_ref].font = label_font
            sheet[cell_ref].font = Font(color=label_font_color)
            # Add left border for A column cells
            if cell_ref.startswith('A'):
                sheet[cell_ref].border = Border(left=medium_border)
        
        # Project summary data with formatting
        data_cells = [
            ("C4", project_object.Project_Name),
            ("C5", project_object.Project_ID),
            ("C6", project_object.Location),
            ("C7", project_object.Post_Code),
            ("C8", project_object.Sector),
            ("C9", project_object.Portfolio_Bus_Unit_Dept_ID),
            ("J4", project_object.Asset_Type),
            ("J5", project_object.Contract_Type),
            ("J6", project_object.Contract_Financial),
            ("J7", project_object.Client),
            ("J8", project_object.Stage_of_Work)
        ]
        
        for cell_ref, value in data_cells:
            sheet[cell_ref] = value if value else ""
            sheet[cell_ref].font = value_font
            sheet[cell_ref].alignment = value_alignment
            # Add borders for specific cells as in reference
            if cell_ref in ["C4", "C5"]:
                if cell_ref == "C4":
                    sheet[cell_ref].border = Border(left=medium_border, top=medium_border)
                else:
                    sheet[cell_ref].border = Border(left=medium_border, right=medium_border, top=medium_border)
        
        # Merge cells for comments (C11:J12)
        sheet.merge_cells('C11:J12')
        sheet["C11"] = project_object.Comments if project_object.Comments else ""
        sheet["C11"].font = value_font
        sheet["C11"].alignment = value_alignment
        
        # Instructions
        sheet["A15"] = "Instructions: Please overwrite all mandatory cells with your own business data"
        sheet["A16"] = "You may also upload this template directly as sample data to experience"
        sheet["A17"] = "Octant AI with sample operational data"
        
        # Merge cells for instructions
        sheet.merge_cells('A14:F14')
        
        # Monthly data headers with formatting
        sheet["F20"] = "MONTHLY DATA"
        sheet["F20"].font = summary_font
        sheet["F20"].fill = summary_fill
        sheet["F20"].alignment = summary_alignment
        
        monthly_headers = [
            ("A21", "Reporting Date*"),
            ("B21", "Approved Budget   (including contingency)"),
            ("C21", "Forecast End Date (practical completion)"),
            ("D21", "Forecast Final Cost* (including  contingency)"),
            ("E21", "Contingency Remaining"),
            ("F21", "Actual Cost to Date*"),
            ("G21", "Forecast Final Revenue*"),
            ("H21", "Actual Revenue to Date*"),
            ("I21", "Notes"),
            ("J21", "Comments")
        ]
        
        for cell_ref, header_text in monthly_headers:
            sheet[cell_ref] = header_text
            sheet[cell_ref].font = monthly_header_font
            sheet[cell_ref].fill = monthly_header_fill
            sheet[cell_ref].alignment = monthly_header_alignment
            
            # Add borders for monthly headers
            col_letter = cell_ref[0]
            if col_letter == 'A':
                sheet[cell_ref].border = Border(left=medium_border, right=thin_border, top=medium_border, bottom=thin_border)
            elif col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                sheet[cell_ref].border = Border(left=thin_border, right=thin_border, top=medium_border, bottom=thin_border)
            else:  # I, J columns
                sheet[cell_ref].border = Border(left=thin_border, top=medium_border, bottom=thin_border)
        
        # Merge cells for Notes and Comments columns (I21:K21)
        sheet.merge_cells('I21:K21')
        
        
        # Monthly data rows
        start_row = 22
        # Define custom currency format
        currency_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        date_format = 'DD/MM/YYYY'  # Date format for reporting date
        for record_idx, monthly_record in enumerate(project_object.Monthly_data):
            row = start_row + record_idx
            
            # Data values
            data_values = [
                (f"A{row}", monthly_record.Date),
                (f"B{row}", monthly_record.approved_budget),
                (f"C{row}", monthly_record.forecast_end_date),
                (f"D{row}", monthly_record.forecast_final_cost),
                (f"E{row}", monthly_record.contingency_remaining),
                (f"F{row}", monthly_record.actual_cost_to_date),
                (f"G{row}", monthly_record.forecast_final_revenue),
                (f"H{row}", monthly_record.actual_revenue_to_date),
                (f"I{row}", monthly_record.notes)
            ]
            
            # Currency columns that need special formatting
            currency_columns = ['B', 'D', 'E', 'F', 'G', 'H']
            date_columns = ['A']  # Reporting Date
            for cell_ref, value in data_values:
                # Convert date string to actual date object for Excel
                if cell_ref.startswith('A') and value and isinstance(value, str):
                    try:
                        # Parse the date string (DD/MM/YYYY format) to a datetime object
                        date_obj = datetime.datetime.strptime(value, "%d/%m/%Y").date()
                        sheet[cell_ref] = date_obj
                    except ValueError:
                        # If parsing fails, keep as string
                        sheet[cell_ref] = value
                else:
                    sheet[cell_ref] = value
                    
                sheet[cell_ref].font = data_font
                
                # Add borders for data cells
                col_letter = cell_ref[0]
                if col_letter == 'A':
                    sheet[cell_ref].border = Border(left=medium_border, right=thin_border, top=thin_border, bottom=thin_border)
                elif col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    sheet[cell_ref].border = data_border
                else:  # I column
                    sheet[cell_ref].border = Border(left=thin_border, top=thin_border, bottom=thin_border, right=medium_border)
                
                # Apply currency format to specific columns AFTER setting value and borders
                if col_letter in currency_columns:
                    sheet[cell_ref].number_format = currency_format
                if col_letter in date_columns:
                    sheet[cell_ref].number_format = date_format
            # Merge cells for Notes column (I{row}:K{row})
            sheet.merge_cells(f'I{row}:K{row}')
    
    # Save the workbook
    try:
        workbook.save(filename)
        print(f"Successfully saved {filename}")
        
    except Exception as e:
        print(f"Error saving {filename}: {e}")
        # Try alternative save method
        try:
            workbook.save(filename.replace('.xlsx', '_backup.xlsx'))
            print(f"Saved backup version: {filename.replace('.xlsx', '_backup.xlsx')}")
        except:
            print(f"Failed to save backup version")
    
    workbook.close()

def validate_excel_formatting(filename):
    """
    Validate that the Excel file has proper formatting applied
    """
    try:
        wb = px.load_workbook(filename)
        if not wb.sheetnames:
            return False, "No sheets found"
        
        sheet = wb[wb.sheetnames[0]]
        
        # Check key formatting elements
        checks = []
        
        # Check header formatting
        a1_cell = sheet["A1"]
        checks.append(("A1 font bold", a1_cell.font and a1_cell.font.bold))
        checks.append(("A1 fill color", a1_cell.fill and hasattr(a1_cell.fill, 'start_color') and a1_cell.fill.start_color.rgb == 'FFD9D9D9'))
        
        # Check main header
        h1_cell = sheet["H1"]
        checks.append(("H1 font size", h1_cell.font and h1_cell.font.size == 18))
        checks.append(("H1 alignment", h1_cell.alignment and h1_cell.alignment.horizontal == 'center'))
        
        # Check project summary
        f2_cell = sheet["F2"]
        checks.append(("F2 fill color", f2_cell.fill and hasattr(f2_cell.fill, 'start_color') and f2_cell.fill.start_color.rgb == 'FFAFE7E7'))
        
        # Check borders
        a21_cell = sheet["A21"]
        checks.append(("A21 borders", a21_cell.border and a21_cell.border.left and a21_cell.border.left.style == 'medium'))
        
        # Check merged cells
        merged_count = len(sheet.merged_cells.ranges)
        checks.append(("Merged cells", merged_count > 0))
        
        # Report results
        passed = sum(1 for _, check in checks if check)
        total = len(checks)
        
        wb.close()
        
        success = passed == total
        message = f"Validation: {passed}/{total} checks passed"
        
        if not success:
            failed_checks = [name for name, check in checks if not check]
            message += f". Failed: {', '.join(failed_checks)}"
        
        return success, message
        
    except Exception as e:
        return False, f"Error validating file: {e}"

# def debug_dataframe_structure(df):
#     """Debug function to analyze DataFrame structure and find percentage columns"""
#     st.write("=== DEBUGGING DATAFRAME STRUCTURE ===")
#     st.write(f"DataFrame shape: {df.shape}")
#     st.write(f"Column count: {len(df.columns)}")
    
#     # Show first few columns with their indices
#     st.write("Column indices and names:")
#     for i, col in enumerate(df.columns[:20]):  # Show first 20 columns
#         st.write(f"  {i}: '{col}'")
    
#     # Look for Complete columns specifically
#     st.write("\nColumns containing 'complete':")
#     complete_cols = []
#     for i, col in enumerate(df.columns):
#         if "complete" in str(col).lower():
#             complete_cols.append((i, col))
#             st.write(f"  {i}: '{col}'")
    
#     if complete_cols:
#         # Show sample data from percentage columns
#         for col_idx, col_name in complete_cols[:3]:  # Show first 3 matching columns
#             st.write(f"\nSample data from column '{col_name}':")
#             sample_data = df[col_name].dropna().head(10)
#             st.write(sample_data.tolist())
#             st.write(f"Data type: {df[col_name].dtype}")
#     else:
#         st.write("No columns found containing 'complete'")
    
#     # Show project column data
#     if pf.PROJECT_NAME < len(df.columns):
#         project_col = df.columns[pf.PROJECT_NAME]
#         st.write(f"\nProject column (index {pf.PROJECT_NAME}): '{project_col}'")
#         unique_projects = df[project_col].dropna().unique()
#         st.write(f"Unique projects ({len(unique_projects)}): {unique_projects[:5].tolist()}")
    
#     st.write("=== END DEBUGGING ===")