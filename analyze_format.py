import openpyxl as px
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd

# Load the reference file
ref_file = 'Fredon calibration data.xlsx'
try:
    workbook = px.load_workbook(ref_file)
    print('Available sheets in reference file:')
    for sheet in workbook.sheetnames:
        print(f'  - {sheet}')
    
    # Examine the first sheet (likely the template)
    first_sheet = workbook[workbook.sheetnames[0]]
    print(f'\nAnalyzing sheet: {first_sheet.title}')
    print(f'Sheet dimensions: {first_sheet.max_row} rows x {first_sheet.max_column} columns')
    
    # Check key cells for formatting
    print('\nSample cell formatting:')
    sample_cells = ['A1', 'A4', 'H1', 'F2', 'A21', 'B21', 'C4', 'J4']
    for cell_ref in sample_cells:
        cell = first_sheet[cell_ref]
        if cell.value:
            print(f'Cell {cell_ref}: "{cell.value}"')
            if cell.font:
                print(f'  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}')
            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                print(f'  Fill: {cell.fill.start_color.rgb}')
            if cell.alignment:
                print(f'  Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}')
            print()
    
    # Check column widths
    print('Column widths:')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        width = first_sheet.column_dimensions[col].width
        if width:
            print(f'  Column {col}: {width}')
            
except Exception as e:
    print(f'Error: {e}')
