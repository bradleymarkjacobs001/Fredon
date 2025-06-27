"""
Comprehensive comparison between reference and generated files
"""
import openpyxl as px
import os

def compare_files():
    """Compare reference template with generated file"""
    reference_file = "Fredon calibration data.xlsx"
    generated_file = "portfolio_output_Calibrate.xlsx"
    
    if not os.path.exists(reference_file):
        print(f"Reference file {reference_file} not found")
        return
    
    if not os.path.exists(generated_file):
        print(f"Generated file {generated_file} not found")
        return
    
    print("=== COMPARING REFERENCE VS GENERATED ===")
    
    # Load both files
    ref_wb = px.load_workbook(reference_file)
    gen_wb = px.load_workbook(generated_file)
    
    ref_sheet = ref_wb[ref_wb.sheetnames[0]]
    gen_sheet = gen_wb[gen_wb.sheetnames[0]]
    
    print(f"Reference sheet: {ref_sheet.title}")
    print(f"Generated sheet: {gen_sheet.title}")
    
    # Compare key formatting cells
    test_cells = [
        "A1", "B1", "H1", "F2", "A4", "C4", "A5", "C5", 
        "F20", "A21", "B21", "C21", "A22", "B22"
    ]
    
    print("\n=== CELL-BY-CELL COMPARISON ===")
    
    for cell_ref in test_cells:
        print(f"\n--- {cell_ref} ---")
        ref_cell = ref_sheet[cell_ref]
        gen_cell = gen_sheet[cell_ref]
        
        # Compare values
        print(f"Value - Ref: '{ref_cell.value}' | Gen: '{gen_cell.value}'")
        
        # Compare fonts
        ref_font_info = f"{ref_cell.font.name}, {ref_cell.font.size}, Bold:{ref_cell.font.bold}" if ref_cell.font else "None"
        gen_font_info = f"{gen_cell.font.name}, {gen_cell.font.size}, Bold:{gen_cell.font.bold}" if gen_cell.font else "None"
        print(f"Font - Ref: {ref_font_info} | Gen: {gen_font_info}")
        match_font = ref_font_info == gen_font_info
        print(f"Font Match: {match_font}")
        
        # Compare fills
        ref_fill = ref_cell.fill.start_color.rgb if ref_cell.fill and hasattr(ref_cell.fill, 'start_color') and ref_cell.fill.start_color else "None"
        gen_fill = gen_cell.fill.start_color.rgb if gen_cell.fill and hasattr(gen_cell.fill, 'start_color') and gen_cell.fill.start_color else "None"
        print(f"Fill - Ref: {ref_fill} | Gen: {gen_fill}")
        match_fill = ref_fill == gen_fill
        print(f"Fill Match: {match_fill}")
        
        # Compare alignment
        ref_align = f"H:{ref_cell.alignment.horizontal}, V:{ref_cell.alignment.vertical}" if ref_cell.alignment else "None"
        gen_align = f"H:{gen_cell.alignment.horizontal}, V:{gen_cell.alignment.vertical}" if gen_cell.alignment else "None"
        print(f"Alignment - Ref: {ref_align} | Gen: {gen_align}")
        match_align = ref_align == gen_align
        print(f"Alignment Match: {match_align}")
        
        # Overall match
        overall_match = match_font and match_fill and match_align
        print(f"Overall Match: {'✅' if overall_match else '❌'}")
    
    # Check column widths
    print(f"\n=== COLUMN WIDTH COMPARISON ===")
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for col in columns:
        ref_width = ref_sheet.column_dimensions[col].width
        gen_width = gen_sheet.column_dimensions[col].width
        match = abs(ref_width - gen_width) < 0.1  # Allow small differences
        print(f"Column {col} - Ref: {ref_width:.2f} | Gen: {gen_width:.2f} | Match: {'✅' if match else '❌'}")
    
    print(f"\n=== ANALYSIS COMPLETE ===")

if __name__ == "__main__":
    compare_files()
