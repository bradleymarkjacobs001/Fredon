"""
Analyze the generated Excel file to identify formatting issues
"""
import openpyxl as px
import os

def analyze_generated_file():
    """Analyze the generated portfolio file"""
    try:
        filename = "portfolio_output_Calibrate.xlsx"
        if not os.path.exists(filename):
            print(f"File {filename} not found")
            return
        
        print(f"=== ANALYZING GENERATED FILE: {filename} ===")
        wb = px.load_workbook(filename)
        
        print(f"Available sheets: {wb.sheetnames}")
        
        if not wb.sheetnames:
            print("No sheets found in workbook")
            return
        
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        print(f"Sheet: {sheet_name}")
        print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        
        # Check key cells for formatting
        key_cells = [
            ("A1", "Template Version"),
            ("B1", "Version Value"),
            ("H1", "Main Header"),
            ("F2", "Project Summary"),
            ("A4", "Project Name Label"),
            ("C4", "Project Name Value"),
            ("A21", "Reporting Date Header"),
            ("B21", "Budget Header"),
            ("A22", "First Data Row"),
        ]
        
        print("\n=== CELL ANALYSIS ===")
        for cell_ref, description in key_cells:
            try:
                cell = sheet[cell_ref]
                print(f"\n{description} ({cell_ref}):")
                print(f"  Value: '{cell.value}'")
                
                if cell.font:
                    print(f"  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")
                else:
                    print("  Font: None")
                
                if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                    color = cell.fill.start_color.rgb if cell.fill.start_color.rgb != '00000000' else "None"
                    print(f"  Fill: {color}")
                else:
                    print("  Fill: None")
                
                if cell.alignment and (cell.alignment.horizontal or cell.alignment.vertical):
                    print(f"  Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}")
                else:
                    print("  Alignment: None")
                
                if cell.border:
                    borders = []
                    if cell.border.left and cell.border.left.style:
                        borders.append(f"L={cell.border.left.style}")
                    if cell.border.right and cell.border.right.style:
                        borders.append(f"R={cell.border.right.style}")
                    if cell.border.top and cell.border.top.style:
                        borders.append(f"T={cell.border.top.style}")
                    if cell.border.bottom and cell.border.bottom.style:
                        borders.append(f"B={cell.border.bottom.style}")
                    
                    if borders:
                        print(f"  Border: {', '.join(borders)}")
                    else:
                        print("  Border: None")
                else:
                    print("  Border: None")
                    
            except Exception as e:
                print(f"Error analyzing cell {cell_ref}: {e}")
        
        # Check merged cells
        print(f"\n=== MERGED CELLS ===")
        if sheet.merged_cells.ranges:
            for merged_range in sheet.merged_cells.ranges:
                print(f"  {merged_range}")
        else:
            print("  No merged cells found")
        
        print(f"\n=== END ANALYSIS ===")
        
    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_generated_file()
