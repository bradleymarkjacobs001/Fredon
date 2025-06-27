"""
Enhanced formatting analysis to compare reference template with our output
"""
import openpyxl as px
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

def analyze_reference_formatting():
    """Analyze reference file formatting in detail"""
    try:
        reference_file = "Fredon calibration data.xlsx"
        if not os.path.exists(reference_file):
            print(f"Reference file {reference_file} not found")
            return
        
        wb = px.load_workbook(reference_file)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        print(f"=== DETAILED REFERENCE FORMATTING ANALYSIS ===")
        print(f"Sheet: {sheet_name}")
        print(f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        
        # Analyze specific key cells for formatting
        key_cells = [
            ("A1", "Template version cell"),
            ("B1", "Version value cell"),
            ("H1", "Main header cell"),
            ("F2", "Project summary header"),
            ("A4", "Project name label"),
            ("C4", "Project name value"),
            ("A5", "Project ID label"),
            ("C5", "Project ID value"),
            ("F20", "Monthly data header"),
            ("A21", "Reporting date header"),
            ("B21", "Budget header"),
            ("A22", "First data row - date"),
            ("B22", "First data row - budget"),
        ]
        
        for cell_ref, description in key_cells:
            cell = sheet[cell_ref]
            print(f"\n{description} ({cell_ref}):")
            print(f"  Value: '{cell.value}'")
            if cell.font:
                print(f"  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")
            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                print(f"  Fill: {cell.fill.start_color.rgb}")
            if cell.alignment:
                print(f"  Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}")
            if cell.border and any([cell.border.left.style, cell.border.right.style, 
                                   cell.border.top.style, cell.border.bottom.style]):
                print(f"  Border: L={cell.border.left.style}, R={cell.border.right.style}, "
                      f"T={cell.border.top.style}, B={cell.border.bottom.style}")
        
        # Check for merged cells
        print(f"\nMerged cells:")
        for merged_range in sheet.merged_cells.ranges:
            print(f"  {merged_range}")
        
        # Analyze borders in detail
        print(f"\nBorder analysis:")
        border_count = 0
        for row in range(1, min(25, sheet.max_row + 1)):
            for col in range(1, min(11, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.border and any([cell.border.left.style, cell.border.right.style, 
                                       cell.border.top.style, cell.border.bottom.style]):
                    border_count += 1
                    if border_count <= 5:  # Show first 5 bordered cells
                        print(f"  {px.utils.get_column_letter(col)}{row}: "
                              f"L={cell.border.left.style}, R={cell.border.right.style}, "
                              f"T={cell.border.top.style}, B={cell.border.bottom.style}")
        
        if border_count > 5:
            print(f"  ... and {border_count - 5} more cells with borders")
        
        print(f"\n=== END DETAILED ANALYSIS ===")
        
    except Exception as e:
        print(f"Error analyzing reference file: {e}")

def test_current_output():
    """Test the current Excel export function"""
    try:
        # Import our modules
        import sys
        sys.path.append('.')
        
        from Fredon_Methods_test import create_data_objects
        
        # Load some test data
        df_path = "Fredon live data running edit.xlsx"
        if not os.path.exists(df_path):
            print(f"Test data file {df_path} not found")
            return
        
        import pandas as pd
        df = pd.read_excel(df_path)
        
        # Create test objects and export
        portfolio = create_data_objects(df)
        
        if portfolio and portfolio.Projects:
            print(f"\nCreated portfolio with {len(portfolio.Projects)} project(s)")
            
            # Test the Excel export with first project only
            test_portfolio = type(portfolio)()  # Create empty portfolio of same type
            test_portfolio.Projects = portfolio.Projects[:1]  # Just first project
            
            from Fredon_Methods_test import create_excel_file_with_portfolio_data
            
            files_created = create_excel_file_with_portfolio_data(test_portfolio, "test_output")
            print(f"Created files: {files_created}")
            
            # Analyze the first created file
            if files_created:
                analyze_created_file(files_created[0])
        else:
            print("No portfolio data created")
            
    except Exception as e:
        print(f"Error testing current output: {e}")
        import traceback
        traceback.print_exc()

def analyze_created_file(filename):
    """Analyze a file we just created"""
    try:
        print(f"\n=== ANALYZING CREATED FILE: {filename} ===")
        wb = px.load_workbook(filename)
        sheet_name = wb.sheetnames[0]
        sheet = wb[sheet_name]
        
        print(f"Sheet: {sheet_name}")
        
        # Check key cells
        key_cells = ["A1", "H1", "F2", "A4", "C4", "F20", "A21"]
        
        for cell_ref in key_cells:
            cell = sheet[cell_ref]
            print(f"{cell_ref}: '{cell.value}'")
            if cell.font:
                print(f"  Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")
            if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                print(f"  Fill: {cell.fill.start_color.rgb}")
        
        print(f"=== END ANALYSIS OF CREATED FILE ===")
        
    except Exception as e:
        print(f"Error analyzing created file: {e}")

if __name__ == "__main__":
    analyze_reference_formatting()
    test_current_output()
