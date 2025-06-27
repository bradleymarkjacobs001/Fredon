"""
Simple tool to verify Excel file formatting
Usage: python verify_excel_format.py [filename]
"""
import sys
import openpyxl as px

def verify_excel_format(filename):
    """Verify that an Excel file has proper formatting"""
    try:
        print(f"Verifying formatting for: {filename}")
        wb = px.load_workbook(filename)
        
        if not wb.sheetnames:
            print("❌ No sheets found in workbook")
            return False
        
        sheet = wb[wb.sheetnames[0]]
        print(f"✅ Checking sheet: {sheet.title}")
        
        # Verification checks
        checks = {
            "Template Version Bold": lambda: sheet["A1"].font and sheet["A1"].font.bold,
            "Template Version Fill": lambda: sheet["A1"].fill and hasattr(sheet["A1"].fill, 'start_color') and sheet["A1"].fill.start_color.rgb == 'FFD9D9D9',
            "Main Header Size": lambda: sheet["H1"].font and sheet["H1"].font.size == 18,
            "Main Header Alignment": lambda: sheet["H1"].alignment and sheet["H1"].alignment.horizontal == 'center',
            "Project Summary Fill": lambda: sheet["F2"].fill and hasattr(sheet["F2"].fill, 'start_color') and sheet["F2"].fill.start_color.rgb == 'FFAFE7E7',
            "Monthly Header Fill": lambda: sheet["A21"].fill and hasattr(sheet["A21"].fill, 'start_color') and sheet["A21"].fill.start_color.rgb == 'FFD9D9D9',
            "Borders Present": lambda: sheet["A21"].border and sheet["A21"].border.left and sheet["A21"].border.left.style == 'medium',
            "Merged Cells": lambda: len(sheet.merged_cells.ranges) > 0,
            "Column Widths": lambda: sheet.column_dimensions['A'].width == 12.55 and sheet.column_dimensions['J'].width == 31.44
        }
        
        print("\n=== FORMATTING VERIFICATION ===")
        passed = 0
        total = len(checks)
        
        for check_name, check_func in checks.items():
            try:
                result = check_func()
                status = "✅" if result else "❌"
                print(f"{status} {check_name}")
                if result:
                    passed += 1
            except Exception as e:
                print(f"❌ {check_name} (Error: {e})")
        
        print(f"\n=== SUMMARY ===")
        print(f"Passed: {passed}/{total} checks")
        
        if passed == total:
            print("🎉 All formatting checks passed! The file is properly formatted.")
        else:
            print(f"⚠️  {total - passed} formatting issues found.")
        
        # Show sample cell values
        print(f"\n=== SAMPLE VALUES ===")
        sample_cells = ["A1", "H1", "F2", "C4", "C5"]
        for cell_ref in sample_cells:
            cell = sheet[cell_ref]
            print(f"{cell_ref}: '{cell.value}'")
        
        wb.close()
        return passed == total
        
    except Exception as e:
        print(f"❌ Error verifying file: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    else:
        filename = "portfolio_output_Calibrate.xlsx"
    
    verify_excel_format(filename)
